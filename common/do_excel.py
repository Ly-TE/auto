# -*- coding: utf-8 -*-
# author = 'API'

from common.do_log import Log
from common.project_path import *
from openpyxl import load_workbook
from common.config import DoConfig


#IP和端口
url=DoConfig(conf_path).get_value('URL','url')
#用例配置项
case_config=DoConfig(conf_path).get_value('CASE_CONFIG','case_config')
case_id=DoConfig(conf_path).get_value('CASE_CONFIG','case_id')

class Case:
    def __init__(self):
        self.id=None
        self.title=None
        self.method=None
        self.url=None
        self.data=None
        self.expected=None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        try:
            lw=load_workbook(file_name)
            self.sheet=lw.get_sheet_by_name(sheet_name)#定位到表单
        except Exception as e:
            Log.error('打开文件，定位表单失败：{}'.format(e))
            raise e

    def get_value(self):
        cases=[]
        for i in range(2,self.sheet.max_row+1):
            case=Case()
            case.id=self.sheet.cell(i,1).value
            case.title=self.sheet.cell(i,2).value
            case.method=self.sheet.cell(i,3).value
            case.url=self.sheet.cell(i,4).value
            case.data=self.sheet.cell(i,5).value
            case.expected=self.sheet.cell(i,6).value
            cases.append(case)

        final_case=[]
        if case_config=='all': #全部用例
            final_case=cases
            Log.debug('执行模块下所有用例')
        else:
            for id in eval(case_id): #用例id参数是一个列表
                for i in cases:
                    if id==int(i.__dict__['id']):
                        final_case.append(i)
                        Log.debug('执行模块下配置的用例')
        return final_case

    #将测试结果写回Excel待补充
    def write_excel(self):
        pass

if __name__=='__main__':
    print(url)
    t=DoExcel(data_path,'boss')

    cases=t.get_value()
    print(cases)
    for c in cases:
        print(c.__dict__) #直接调用每一个object属性
        # print(c.id)
        # print(c.url)
        print(c.data)






